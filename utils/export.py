import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


COLOUR = {
    "header_bg"      : "1A1A2E",   # dark navy
    "header_fg"      : "FFFFFF",   # white text
    "answered_bg"    : "D4F5E2",   # soft green
    "not_found_bg"   : "FDE8CC",   # soft orange
    "manual_bg"      : "E8F0FE",   # soft blue
    "alt_row"        : "F8F9FA",   # light grey alternate row
    "border"         : "CCCCCC",   # light border
    "title_bg"       : "16213E",   # deeper navy for title row
    "title_fg"       : "E2C77A",   # gold accent for title
}



def _thin_border() -> Border:
    thin = Side(style="thin", color=COLOUR["border"])
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_header_row(ws: Worksheet, columns: list[str]):
  
    header_fill   = PatternFill("solid", fgColor=COLOUR["header_bg"])
    header_font   = Font(bold=True, color=COLOUR["header_fg"], size=10)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border        = _thin_border()

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = border

def _write_title_row(ws: Worksheet, num_columns: int, project_name: str = "RFP Responses"):
   
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value     = f"Venuebase — {project_name}"
    title_cell.fill      = PatternFill("solid", fgColor=COLOUR["title_bg"])
    title_cell.font      = Font(bold=True, color=COLOUR["title_fg"], size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _row_fill(status: str, row_idx: int) -> PatternFill:
 
    if status == "answered":
        return PatternFill("solid", fgColor=COLOUR["answered_bg"])
    elif status == "not_found":
        return PatternFill("solid", fgColor=COLOUR["not_found_bg"])
    elif status == "manual":
        return PatternFill("solid", fgColor=COLOUR["manual_bg"])
    else:
   
        colour = COLOUR["alt_row"] if row_idx % 2 == 0 else "FFFFFF"
        return PatternFill("solid", fgColor=colour)


def prepare_xlsx(
    df: pd.DataFrame,
    project_name: str = "RFP Responses",
) -> bytes:

    export_df = df.drop(columns=["Regenerate"], errors="ignore").copy()

    if "Confidence" in export_df.columns:
        export_df["Confidence"] = export_df["Confidence"].apply(
            lambda x: f"{x * 100:.0f}%" if pd.notna(x) else "N/A"
        )

    columns    = list(export_df.columns)
    num_cols   = len(columns)

    col_widths = {
        "Q#"        : 6,
        "Question"  : 55,
        "AI Answer" : 65,
        "Citation"  : 35,
        "Status"    : 14,
        "Confidence": 14,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "RFP Answers"


    _write_title_row(ws, num_columns=num_cols, project_name=project_name)

    _write_header_row(ws, columns)
    ws.row_dimensions[2].height = 22

    wrap_align   = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")
    border       = _thin_border()

    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=3):
        status = str(df.at[row.name, "Status"]) if "Status" in df.columns else ""
        fill   = _row_fill(status, row_idx)

        for col_idx, col_name in enumerate(columns, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(size=10)

            if col_name in ("Q#", "Status", "Confidence"):
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align


        max_lines = 1
        for col_name in ("Question", "AI Answer", "Citation"):
            if col_name in export_df.columns:
                text  = str(row.get(col_name, ""))
                lines = max(1, len(text) // 60 + text.count("\n") + 1)
                max_lines = max(max_lines, lines)

        ws.row_dimensions[row_idx].height = max(18, min(max_lines * 15, 120))

    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        width  = col_widths.get(col_name, 20)
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A3"

    _write_legend_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


#  LEGEND SHEET

def _write_legend_sheet(wb: Workbook):
    
    ws = wb.create_sheet(title="Legend")

    legend_data = [
        ("Colour",  "Status",               "Meaning"),
        ("Green",   "answered",              "AI found a direct answer in the reference documents."),
        ("Orange",  "not_found",             "Answer could not be found in any reference document. Manual review required."),
        ("Blue",    "manual",                "Answer was manually entered or edited by the Venuebase team."),
    ]

    fills = {
        "Green"  : PatternFill("solid", fgColor=COLOUR["answered_bg"]),
        "Orange" : PatternFill("solid", fgColor=COLOUR["not_found_bg"]),
        "Blue"   : PatternFill("solid", fgColor=COLOUR["manual_bg"]),
    }

    header_fill = PatternFill("solid", fgColor=COLOUR["header_bg"])
    header_font = Font(bold=True, color=COLOUR["header_fg"], size=10)
    border      = _thin_border()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70

    for row_idx, (colour_label, status, meaning) in enumerate(legend_data, start=1):
        for col_idx, value in enumerate([colour_label, status, meaning], start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.font   = Font(size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif colour_label in fills:
                cell.fill = fills[colour_label]

        ws.row_dimensions[row_idx].height = 22

    ws.insert_rows(1)
    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value="RFP Answer Legend — Venuebase")
    title.fill      = PatternFill("solid", fgColor=COLOUR["title_bg"])
    title.font      = Font(bold=True, color=COLOUR["title_fg"], size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26